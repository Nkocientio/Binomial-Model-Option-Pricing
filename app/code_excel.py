from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font

def summary(S0, K, T, r,q, sigma, optionType, style, steps, upFactor, uProb, price,color1,color2):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    optionInputs = [['Net Asset Price','Strike Price(K)','Maturity (Years)','Interest Rate(p.a)','Div yield(p.a)','Volatility (p.a)','Option Type','Exercise Style','Steps'],[S0,K,T,r,q,sigma,optionType,style,steps]]
    optionOutputs = [['Upside-Factor(u)','Downside-Factor(d)','Probability (p)','Downside-Prob(1-p)',f'Value of {style} {optionType}'],[upFactor,1.0/upFactor,uProb,1.0-uProb,price]]
    
    ws['B2'] = 'Inputs'
    ws['B2'].fill = color1
    ws['C2'].fill = color2
    for i in range(3,12):
        ws[f'B{i}'] = optionInputs[0][i-3]
        ws[f'B{i}'].fill = color2
        ws[f'C{i}'] = optionInputs[1][i-3]
        ws[f'C{i}'].fill = color1

    ws['B14'] = 'Outputs'
    ws['B14'].fill = color1
    ws['C14'].fill = color2
    for i in range(15,20):
        ws[f'B{i}'] = optionOutputs[0][i-15]
        ws[f'B{i}'].fill = color2
        ws[f'C{i}'] = optionOutputs[1][i-15]
        ws[f'C{i}'].fill = color1

    ws.column_dimensions['B'].width = 19
    return wb

def BermudanDates(wb,style, dates):
    r = 7 if style == 'Compound' else 0
    ws = wb['Binomial Tree']

    ws[f'A{8+r}'] = 'Can Exercise on'
    ws[f'A{8+r}'].font = Font(bold=True)
    ws[f'A{8+r}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    for i in range(len(dates)):
        ws[f"A{9+r+i}"] = dates[i]
        ws[f"A{9+r+i}"].fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        
    ws.column_dimensions['A'].width = 13
    return wb

def number_to_letter(n):
    if n <= 0:
        n=1
    letters = []
    while n:
        n, remainder = divmod(n - 1, 26)
        letters.append(chr(65 + remainder))
    return ''.join(reversed(letters))

def print_excel(style, optionType, tree, isExercised, steps, upFactor, uProb, strikes, anotherStyle, dividends, dt, onCallPut, compoundStrike, compoundStep, avgType, avgWhat, S0, K, T, r,q, sigma, price, exerciseDates, divType):
    
    optionC = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    assetC = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    greyC = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")

    # Summary
    wb = summary(S0, K, T, r,q, sigma, optionType, style, steps, upFactor, uProb, price,assetC,greyC)
    ws = wb.create_sheet('Binomial Tree')
    
    startR, startC = 4, 3
    colShifter = 0
    underlying = "Asset"

    if style == 'Asian':
        ws['A1'] = f'{avgType} {avgWhat} Averaging'
        ws['A1'].font = Font(italic=True, size=10)
    elif style == 'Compound':
        results = [[f'{optionType} on','K_c','T1 (year)','Style'],[onCallPut,compoundStrike,compoundStep * dt,anotherStyle]]
        for i in range(8,12):
            ws[f'A{i}'] = results[0][i-8]
            ws[f'A{i}'].fill = greyC
            ws[f'B{i}'] = results[1][i-8]
            ws[f'B{i}'].fill = assetC
        ws['A8'].fill = assetC
        ws['B8'].fill = assetC
        
        ws['A13'] = "Blue: exercised Compound"
        ws['A13'].font =  Font(color="00B0F0", size=11)
        startC +=1
        colShifter = 1
        underlying = f'{underlying}(option)'

    if style == 'Bermudan' or anotherStyle == 'Bermudan':
        wb = BermudanDates(wb,style, exerciseDates)
        ws = wb['Binomial Tree']
        
    rowRef = steps
    for j in range(steps + 1):
        col = number_to_letter(j+startC)
        exerciseColor = '00B0F0' if (style == 'Compound' and (j <= compoundStep)) else 'FF0000'
        for i in range(j+1):
            # Asset
            ws[f'{col}{2*i+startR+rowRef}'] = tree[2*i+rowRef,j]
            ws[f'{col}{2*i+startR+rowRef}'].fill = assetC
            
            # Option
            ws[f'{col}{2*i+startR+rowRef+1}'] = tree[2*i+rowRef+1,j]
            ws[f'{col}{2*i+startR+rowRef+1}'].fill = optionC
            if isExercised[2*i + rowRef+1, j]:
                ws[f'{col}{2*i+startR+rowRef+1}'].font = Font(color=exerciseColor)

        rowRef -= 1
        # Year row
        ws[f'{col}{startR + 3 + 2 * steps}'] = j*dt
        ws[f'{col}{startR + 3 + 2 * steps}'].fill = greyC

        # Asian Strikes
        if style == "Asian" and avgWhat == "Strike":
            ws[f'{col}{startR + 5 + 2 * steps}'] = strikes[j]
            ws[f'{col}{startR + 5 + 2 * steps}'].fill = greyC

        # Cash Divs
        if divType == "Cash":
            ws[f'{col}{2}'] = dividends[j]
            ws[f'{col}{1}'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            ws[f'{col}{2}'].fill = greyC

    if style == 'Compound':
        ws[f'{number_to_letter(startC + compoundStep)}{startR + 3 + 2 * steps}'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    elif style == "Asian" and avgWhat == "Strike":
        ws[f'{number_to_letter(startC-1)}{startR + 5 + 2 * steps}'] = "Strike(K)"
        ws[f'{number_to_letter(startC-1)}{startR + 5 + 2 * steps}'].fill = assetC
        
    ws[ f'{number_to_letter(2 + colShifter)}{2 * steps + startR+3}' ] = "Year"
    ws[ f'{number_to_letter(2 + colShifter)}{2 * steps + startR+3}' ].fill = assetC

    # Headings
    ws['A2'] = f'{style} {anotherStyle} {optionType}' 
    ws['A2'].font = Font(color="70AD47")
    if style == "Compound":
        optionType = onCallPut
    ws['A3'] = f'Red: Exercised {optionType}'
    ws['A3'].font = Font(color="FF0000")
    ws['A4']= underlying
    ws['A4'].fill = assetC
    ws['A5'] = 'Option'
    ws['A5'].fill = optionC

    if divType == "Cash":
        ws[f'{number_to_letter(startC)}{1}'] = "PV"
        ws[f'{number_to_letter(startC)}{1}'].fill = assetC
        ws[f'{number_to_letter(startC)}{2}'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        
        mid = max(1,steps//2)
        ws[f'{number_to_letter(mid+startC)}{1}'] = 'Cash Dividends'

    return wb

def get_workbook_as_bytes(wb):
    with BytesIO() as buffer:
        wb.save(buffer)
        return buffer.getvalue()
